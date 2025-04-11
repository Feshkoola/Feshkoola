import http.server
import socketserver
import os
from datetime import datetime
import json
from openpyxl import Workbook, load_workbook

PORT = 8000
DIRECTORY = "."

class MyHttpRequestHandler(http.server.SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=DIRECTORY, **kwargs)

    def do_POST(self):
        if self.path == '/send-email':
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            self.send_response(200)
            self.send_header('Content-type', 'text/html')
            self.end_headers()
            self.wfile.write(b'Purchase logged')
            log_purchase(post_data)

def log_purchase(file_data):
    log_file = 'purchase_log.xlsx'
    purchase_info = json.loads(file_data.decode('utf-8'))
    username = purchase_info['Username']
    address = purchase_info['Address']
    total_cost = purchase_info['TotalCost']
    products = purchase_info['Products']

    if os.path.exists(log_file):
        wb = load_workbook(log_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(['Timestamp', 'Username', 'Address', 'Total Cost', 'Product Name', 'Quantity', 'Price'])

    for product in products:
        ws.append([
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            username,
            address,
            total_cost,
            product['ProductName'],
            product['Quantity'],
            product['Price']
        ])

    wb.save(log_file)

handler = MyHttpRequestHandler

with socketserver.TCPServer(("0.0.0.0", PORT), handler) as httpd:
    print(f"Serving at port {PORT}")
    httpd.serve_forever()